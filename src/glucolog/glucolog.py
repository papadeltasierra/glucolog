"""Read a GlucoLog backup database."""
import os
import sys
import re
import glob
import argparse
from argparse import Namespace
import datetime
import calendar
from logging import (
    getLogger,
    StreamHandler,
    DEBUG,
    INFO,
    WARNING,
    Formatter,
    FileHandler,
)
from typing import List
import sqlite3
from sqlite3 import Cursor
import csv
import yaml
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.worksheet import Worksheet

PROC_NAME = "glucolog"

RGX_SAFE_SQL_NAME = re.compile(r"^[a-z_][a-z0-9_@$]*$", re.IGNORECASE)

EXCEL_EPOCH = 25569
DAY_IN_SECS = 24 * 60 * 60
LANG_FILE_GLOB = "lang_??.yml"
FMT_LANG_FILE = "lang_{language}.yml"

# Output file formats.
CSV = "csv"
EXCEL = "excel"
FORMAT_CHOICES = [CSV, EXCEL]
FORMAT_SUFFIXES = {CSV: ".csv", EXCEL: ".xlsx"}


# Translation file keys.
TABLES = "tables"
COLUMNS = "columns"

DEFAULT_WORKSHEET = "Sheet"

CSV_SEPARATOR = ["============"] * 5

# Formatting dict indexes.
FUNC = "function"
STYLE = "style"
WIDTH = "width"
DATA = "data"

# Errors.
ERR_NO_SUCH_TABLE = "no such table"
ERR_NO_SUCH_COLUMN = "no such column"


def entry_exit(func):
    """Decorate a function with entry and exit tracing."""
    func_name = func.__name__

    def _entry_exit(*args, **kwargs):
        log.debug("Entry: { %s", func_name)
        rsp = func(*args, **kwargs)
        log.debug("Exit: } %s", func_name)
        return rsp

    return _entry_exit


# We have to define reformatting functions before the definition of
# the reformatting.
@entry_exit
def day_month_year(excel, value):
    """Parse a day-month-year datestamp."""
    datestamp = datetime.datetime.strptime(value, "%d-%m-%Y")
    if not excel:
        datestamp = datestamp.strftime("%Y-%m-%d")

    return datestamp


@entry_exit
def hour_minute(excel, value):
    """Parse an hour-minute timestamp."""
    # The data can contain "24:00" which is not a valid time!
    if value == "24:00":
        value = "00:00"
    timestamp = datetime.datetime.strptime(value, "%H:%M")
    # if excel:
    #    timestamp = time_to_days(timestamp)
    # else:
    if not excel:
        timestamp = timestamp.strftime("%H:%M:00")

    return timestamp


@entry_exit
def unix_date_microseconds(excel, value):
    """Parse a unix timestamp with microseconds."""
    datestamp = datetime.datetime.fromtimestamp(int(value / 1000))
    # if excel:
    #    datestamp = to_excel(datestamp)
    # else:
    if not excel:
        datestamp = datestamp.strftime("%Y-%m-%d")

    return datestamp


@entry_exit
def time_seconds(excel, value):
    """Parse a timestamp in seconds of the day."""
    if excel:
        # print(value)
        timestamp = datetime.timedelta(seconds=int(value / 1000))
        # timestamp = timedelta_to_days(timestamp)
        # print(timestamp)
        # sys.exit(9)
    else:
        today = datetime.date.timetuple(datetime.date.today())
        today_seconds = calendar.timegm(today)
        today_seconds = today_seconds + int(value / 1000)
        timestamp = datetime.datetime.fromtimestamp(today_seconds)
        timestamp = timestamp.strftime("%H:%M")

    return timestamp


date_style = NamedStyle(name="date", number_format="YYYY-MM-DD")
time_style = NamedStyle(name="time", number_format="HH:MM")

# Identify fields that require special treatment.
REFORMAT_FIELDS = {
    "t_parametri": {
        "data_nascita": {
            FUNC: day_month_year,
            STYLE: date_style,
            WIDTH: 12,
        },
        "digiuno": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "mattino": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "primo_pomeriggio": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "tardo_pomeriggio": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "sera": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "periodocustoms1": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "periodocustoms2": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "periodocustoms3": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
        "periodocustoms4": {
            FUNC: hour_minute,
            STYLE: time_style,
        },
    },
    "t_risultati": {
        "data": {
            FUNC: unix_date_microseconds,
            STYLE: date_style,
            WIDTH: 12,
        },
        "ora": {
            FUNC: time_seconds,
            STYLE: time_style,
        },
        "periodo": {DATA: True},
    },
}

# Get a logging logger.
log = getLogger()


class CsvExport:
    """Class describing creation of a CSV file."""

    @entry_exit
    def __init__(self, filename: str):
        """Create a CSV file."""
        self.file = open(filename, "w", newline="")
        self.csv_writer = csv.writer(self.file)
        self.first_page = True

    @entry_exit
    def worksheet(self, title: str):
        """Fake adding a worksheet to the CSV file."""
        # CSV files don't have the concept of worksheets so we just add all the data to
        # a single file with a simple separator marker between "worksheets".
        if not self.first_page:
            self.csv_writer.writerow(CSV_SEPARATOR)
        self.first_page = False
        self.csv_writer.writerow([title])
        return None

    @entry_exit
    def columns(self, _worksheet: Worksheet, columns: List[str]):
        """Write table column names to the CSV file."""
        self.csv_writer.writerow(columns)

    @entry_exit
    def format_worksheet(
        self, worksheet: Worksheet, _table: str, _columns: List[str]
    ) -> None:
        """No fomratting in a CSV file."""
        pass

    @entry_exit
    def data(self, _worksheet: Worksheet, data: List[str]):
        """Write a row of data to the CSV file."""
        self.csv_writer.writerow(data)

    @entry_exit
    def close(self):
        """Close the CSV file."""
        self.file.close()
        self.file = None
        self.csv_writer = None


class ExcelExport:
    """Class describing creation of an Excel spreadsheet."""

    @entry_exit
    def __init__(self, filename: str):
        """Create an Excel Workbook."""
        self.excel_file = filename
        self.workbook = Workbook()
        self.workbook.iso_dates = True

    @entry_exit
    def worksheet(self, title: str):
        """Create a new Excel worksheet."""
        worksheet = self.workbook.create_sheet(title)
        return worksheet

    @entry_exit
    def columns(self, worksheet: Worksheet, columns: List[str]):
        """Write the column names to the Excel worksheet."""
        worksheet.append(columns)

    @entry_exit
    def format_worksheet(
        self, worksheet: Worksheet, table: str, columns: List[str]
    ) -> None:
        """Set formatting for columns in the worksheet if appropriate."""
        for iindex, (cell,) in enumerate(worksheet.iter_cols(min_row=1, max_row=1)):
            try:
                style = REFORMAT_FIELDS[table][columns[iindex]][STYLE]
                column_width = REFORMAT_FIELDS[table][columns[iindex]][WIDTH]
                log.debug(
                    "applying style '%s' to '%s:%s'.",
                    style.name,
                    table,
                    columns[iindex],
                )
            except KeyError:
                # This just means we don't want to style this field.
                style = None
                column_width = None

            # First set the column width, if present.
            if column_width:
                worksheet.column_dimensions[cell.column_letter].width = column_width
                log.debug("column width set to %d", column_width)
            if style:
                # We have to loop over all cells and apply the style.
                for (cell,) in worksheet.iter_rows(
                    min_row=1, min_col=cell.column, max_col=cell.column
                ):
                    # We don't log here as that creates LOTS of logging.
                    cell.style = style

    @entry_exit
    def data(self, worksheet: Worksheet, data: List[str]):
        """Write a row to the Excel worksheet."""
        worksheet.append(data)

    @entry_exit
    def close(self):
        """Write and close the Excel spreadsheet."""
        # First remove the default "Sheet" worksheet.
        sheet = self.workbook[DEFAULT_WORKSHEET]
        self.workbook.remove(sheet)

        # Now write the workbook.
        self.workbook.save(filename=self.excel_file)
        self.excel_file = None
        self.workbook = None


@entry_exit
def translate_table(args: Namespace, table: str):
    """If possible, translate table name."""
    log.debug("before: %s", table)
    xlat_table = table
    try:
        xlat_table = args.xlat_to[TABLES][table]
    except KeyError:
        log.error("Missing '%s' translation for table '%s'.", args.xlat, table)
        xlat_table = table

    log.debug("after: %s", xlat_table)
    return xlat_table


@entry_exit
def maybe_translate_table(args: Namespace, table: str):
    """If possible, translate table name."""
    log.debug("before: %s", table)
    xlat_table = table
    if args.xlat:
        xlat_table = translate_table(args, table)

    log.debug("after: %s", xlat_table)
    return xlat_table


@entry_exit
def maybe_translate_tables(args: Namespace, tables: List[str]):
    """If possible, translate table names."""
    log.debug("before: %s", tables)
    xlat_tables = list(tables)
    if args.xlat:
        for iindex, table in enumerate(tables):
            xlat_tables[iindex] = translate_table(args, table)

    log.debug("after: %s", xlat_tables)
    return xlat_tables


@entry_exit
def maybe_translate_table_from(args: Namespace, xlat_table: str):
    """Reverse translate a table name."""
    log.debug("xlat_table: %s", xlat_table)
    table = xlat_table
    if args.xlat:
        try:
            # Reverse translation only happens if the user gives is local language
            # values to translate and these MUST match those in the appropriate language
            # file.
            table = args.xlat_from[TABLES][xlat_table]
        except KeyError:
            log.error("Table '%s' is not recognised." % xlat_table)
            sys.exit(2)

    log.debug("table: %s", table)
    return table


@entry_exit
def maybe_translate_columns(args: Namespace, columns: List[str]):
    """If possible, translate column names."""
    log.debug("before: %s", columns)
    xlat_columns = list(columns)
    if args.xlat:
        for iindex, column in enumerate(columns):
            try:
                xlat_columns[iindex] = args.xlat_to[COLUMNS][column]
            except KeyError:
                log.error(
                    "Missing '%s' translation for column '%s'.", args.xlat, column
                )
                xlat_columns[iindex] = column

    log.debug("after: %s", xlat_columns)
    return xlat_columns


@entry_exit
def maybe_translate_columns_from(args: Namespace, columns: List[str]):
    """If possible, translate column names."""
    log.debug("before: %s", columns)
    it_columns = list(columns)
    if args.xlat:
        for iindex, column in enumerate(columns):
            try:
                # Reverse translation only happens if the user gives is local language
                # values to translate and these MUST match those in the appropriate
                # language file.
                it_columns[iindex] = args.xlat_from[COLUMNS][column]
            except KeyError:
                log.error("Column '%s' is not recognised." % column)
                sys.exit(2)

    log.debug("after: %s", it_columns)
    return it_columns


@entry_exit
def maybe_translate_data(
    args: Namespace, it_table, it_columns: List[str], it_data: tuple
):
    """If possible, translate column names."""
    log.debug("before: %s", it_data)
    data = list(it_data)
    if args.xlat:
        for iindex, it_column in enumerate(it_columns):
            try:
                if REFORMAT_FIELDS[it_table][it_column][DATA]:
                    try:
                        # Yes, some data fields are empty!
                        if data[iindex]:
                            data[iindex] = args.xlat_to[DATA][it_data[iindex]]
                    except KeyError:
                        log.error(
                            (
                                "Missing '%s' translation for data '%s' "
                                "from column '%s:%s'."
                            ),
                            args.xlat,
                            it_data[iindex],
                            it_table,
                            it_column,
                        )
                        raise KeyError("Just to trigger arm below...")

            except KeyError:
                # This just means that we don't need a translation of this
                # data field.
                pass

    log.debug("after: %s", data)
    return tuple(data)


@entry_exit
def list_tables(args: Namespace, cur: Cursor):
    """List the tables available in the database."""
    rows = cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
    it_tables = [row[0] for row in rows]

    log.debug("tables: %s", it_tables)
    return it_tables


@entry_exit
def do_list_tables(args: Namespace, cur: Cursor):
    """List the tables available in the database."""
    print("Tables found")
    print("============")
    rows = list_tables(args, cur)
    rows = maybe_translate_tables(args, rows)
    for row in rows:
        print(row)


@entry_exit
def list_columns(args: Namespace, cur: Cursor, it_table: str):
    """List the columns available from the indicated table."""
    log.debug("list columns for table '%s'", it_table)
    # Sanitize the table name to ensure that SQL injextion is not possible.
    assert re.match(r"[a-z][a-z0-9_@$]*", it_table), (
        "'%s' is an invalid table name and could be used for an "
        "SQL injection attack." % it_table
    )
    try:
        rows = cur.execute("SELECT * FROM %s" % it_table)  # nosec
    except sqlite3.OperationalError as ee:
        if ERR_NO_SUCH_TABLE in str(ee):
            log.error("Table '%s' is not recognised.", it_table)
            sys.exit(2)
        else:  # pragma: no cover
            # Some other exception; just raise it.
            raise (ee)

    it_columns = [description[0] for description in rows.description]

    log.debug("columns: %s", it_columns)
    return it_columns


@entry_exit
def do_list_columns(args: Namespace, cur: Cursor):
    """List the columns available from the indicated table."""
    assert "table" in args, "Table name should have been supplied"

    table = maybe_translate_table_from(args, args.table)

    title = "Columns for table '%s'" % args.table
    print(title)
    print("=" * len(title))
    columns = list_columns(args, cur, table)
    columns = maybe_translate_columns(args, columns)
    for column in columns:
        print(column)


def format_data(args: Namespace, it_table: str, it_columns: List[str], it_row: tuple):
    """Reformat the data in rows, typically for time or date formatting."""
    formatted_row = list(it_row)
    for iindex, it_column in enumerate(it_columns):
        try:
            formatted_row[iindex] = REFORMAT_FIELDS[it_table][it_column][FUNC](
                args.format == EXCEL, it_row[iindex]
            )
        except KeyError:
            # Just means there is no reformatter for this field.
            pass

    return tuple(formatted_row)


def export_data(args: Namespace, cur: Cursor, it_table: str, it_columns: List[str]):
    """Write a CSV file that contains the columns from a specific table."""
    # Each row is returned as a tuple...
    assert RGX_SAFE_SQL_NAME.match(it_table), (
        "'%s' is an invalid table name and could be used for an "
        "SQL injection attack." % it_table
    )
    for it_column in it_columns:
        assert RGX_SAFE_SQL_NAME.match(it_column), (
            "'%s' is an invalid column name and could be used for an "
            "SQL injection attack." % it_column
        )
    try:
        it_rows = cur.execute(  # nosec
            "SELECT %s FROM %s" % (",".join(it_columns), it_table)
        )
    except sqlite3.OperationalError as ee:
        if ERR_NO_SUCH_COLUMN in str(ee):
            log.error("One or more columns are not recognised.")
            sys.exit(2)
        else:  # pragma: no cover
            # Some other exception; just raise it.
            raise (ee)

    rows = []
    for it_row in it_rows:
        it_row = format_data(args, it_table, it_columns, it_row)
        rows.append(maybe_translate_data(args, it_table, it_columns, it_row))

    return rows


@entry_exit
def do_export_table(args: Namespace, cur: Cursor):
    """Write a CSV file that contains the columns from a specific table."""
    assert "table" in args, "Table should have been defined"
    assert "format" in args, "Output file formation should have been defined"

    export_file = (
        CsvExport(args.output) if args.format == CSV else ExcelExport(args.output)
    )

    it_table = maybe_translate_table_from(args, args.table)

    # If no columns were specified then we dump all columns, which requires
    # listing them first.
    if not args.columns:
        it_columns = list_columns(args, cur, it_table)
        columns = maybe_translate_columns(args, it_columns)
    else:
        columns = args.columns
        it_columns = maybe_translate_columns_from(args, columns)

    worksheet = export_file.worksheet(args.table)
    export_file.columns(worksheet, columns)
    rows = export_data(args, cur, it_table, it_columns)
    for row in rows:
        export_file.data(worksheet, row)

    export_file.format_worksheet(worksheet, it_table, it_columns)
    export_file.close()


@entry_exit
def do_dump_db(args: Namespace, cur: Cursor):
    """Write a CSV file that contains the columns from a specific table."""
    assert "format" in args, "Output file format should have been defined."
    export_file = (
        CsvExport(args.output) if args.format == CSV else ExcelExport(args.output)
    )
    # First get the tables...
    tables = list_tables(args, cur)

    for table in tables:
        xlat_table = maybe_translate_table(args, table)
        worksheet = export_file.worksheet(xlat_table)

        columns = list_columns(args, cur, table)
        xlat_columns = maybe_translate_columns(args, columns)
        export_file.columns(worksheet, xlat_columns)
        rows = export_data(args, cur, table, columns)
        for row in rows:
            export_file.data(worksheet, row)
        export_file.format_worksheet(worksheet, table, columns)

    export_file.close()


@entry_exit
def setup_logging(args):
    """Set up logging."""
    # Default logging level has to be DEBUG as the filtering is done at Handler level
    # and then global level.
    log.setLevel(DEBUG)

    stream_handler = StreamHandler()
    if args.verbose >= 2:
        stream_handler.setLevel(DEBUG)
    elif args.verbose >= 1:
        stream_handler.setLevel(INFO)
    else:
        stream_handler.setLevel(WARNING)

    file_formatter = Formatter(
        fmt="%(asctime)s %(funcName)-16.16s %(levelname)-9.9s %(message)s",
        datefmt="%H:%M:%S",
    )
    file_handler = FileHandler(args.logfile, mode="w")
    file_handler.setFormatter(file_formatter)
    if args.debug >= 2:
        file_handler.setLevel(DEBUG)
    elif args.debug >= 1:
        file_handler.setLevel(INFO)
    else:
        file_handler.setLevel(WARNING)

    log.addHandler(stream_handler)
    log.addHandler(file_handler)


@entry_exit
def read_language_file(language):
    """Read the language file, validate it and return tranlations."""
    # At some point we might want to use the python-i18n package instead.
    xlat_to = {}
    dirname = os.path.dirname(__file__)
    filename = os.path.join(dirname, FMT_LANG_FILE.format(language=language))
    with open(filename, "r") as source:
        xlat_to = yaml.load(source, Loader=yaml.SafeLoader)

    # Now create the reverse translation table.
    xlat_from = {}
    if xlat_to:
        xlat_from[TABLES] = {y: x for (x, y) in xlat_to[TABLES].items()}
        xlat_from[COLUMNS] = {y: x for (x, y) in xlat_to[COLUMNS].items()}
        xlat_from[DATA] = {y: x for (x, y) in xlat_to[DATA].items()}

    return xlat_to, xlat_from


@entry_exit
def list_languages():
    """List language files so user has options for translation."""
    dirname = os.path.dirname(__file__)
    search_glob = os.path.join(dirname, LANG_FILE_GLOB)
    log.debug("search glob: %s", search_glob)
    lang_files = glob.glob(search_glob)
    languages = [os.path.basename(x)[5:7] for x in lang_files]
    log.debug("languages: %s", str(languages))
    return languages


@entry_exit
def parse_args(argv):
    """Parse command line arguments."""
    # Determine default log file (local file with name of script).
    default_logfile = os.path.splitext(os.path.basename(argv[0]))[0] + ".log"

    # Determine available languages.
    languages = list_languages()

    parser = argparse.ArgumentParser(
        prog=PROC_NAME,
        description="Tool to extract information from a GlucoLog back-up database",
    )
    common_group = parser.add_argument_group(
        title="Common options", description="Options which apply to all commands"
    )
    common_group.add_argument(
        "-v",
        "--verbose",
        action="count",
        help="make the program more verbose",
        default=0,
    )
    common_group.add_argument(
        "-d",
        "--debug",
        action="count",
        help="add debugging information to the log file",
        default=0,
    )
    common_group.add_argument(
        "-l", "--logfile", help="specify log file name", default=default_logfile
    )

    if languages:
        # We have some languages so add the option.
        common_group.add_argument(
            "-x",
            "--xlat",
            type=str,
            choices=languages,
            help="language to translate to",
        )

    common_group.add_argument(
        "database",
        type=str,
        action="store",
        help="name of the GlucoLog backup database",
    )

    subparsers = parser.add_subparsers()
    table_parser = subparsers.add_parser("list-tables")
    table_parser.set_defaults(func=do_list_tables, cmd="list-tables")

    columns_parser = subparsers.add_parser("list-columns")
    columns_parser.add_argument(
        "-t", "--table", required=True, help="name of a specific table in the database"
    )
    columns_parser.set_defaults(func=do_list_columns, cmd="list-columns")

    export_parser = subparsers.add_parser("export-table")
    export_parser.add_argument(
        "-t", "--table", required=True, help="name of a specific table in the database"
    )
    export_parser.add_argument(
        "-c",
        "--columns",
        default=None,
        help="comma seperated names of columns in database table",
    )
    export_parser.add_argument(
        "-f",
        "--format",
        choices=FORMAT_CHOICES,
        required=True,
        help="output to CSV file",
    )
    export_parser.add_argument("output", help="name of destination file")
    export_parser.set_defaults(func=do_export_table, cmd="export-table")

    dump_parser = subparsers.add_parser("dump-db")
    dump_parser.set_defaults(func=do_dump_db, cmd="dump-db")
    dump_parser.add_argument(
        "-f",
        "--format",
        choices=FORMAT_CHOICES,
        required=True,
        help="output to CSV file",
    )
    dump_parser.add_argument("output", help="name of destination file")

    args = parser.parse_args(argv[1:])

    # Now perform additional testing, starting with missing/incorrect parameters.
    if "func" not in args:
        parser.error("No database command was given.")

    # Note that "None" is a valid value meaning "all columns".
    if getattr(args, "columns", None) is not None:
        setattr(args, "columns", args.columns.split(","))

    if "format" in args:
        log.debug("check format, '%s' vs output file, '%s'", args.format, args.output)
        if not args.output.endswith(FORMAT_SUFFIXES[args.format]):
            parser.error(
                "Output '%s' filename '%s' does not end in '%s'"
                % (args.format, args.output, FORMAT_SUFFIXES[args.format])
            )

    # If present, read the language file.
    if getattr(args, "xlat", None) is not None:
        log.info("Reading language file for '%s'...", args.xlat)
        xlat_to, xlat_from = read_language_file(args.xlat)
        setattr(args, "xlat_to", xlat_to)
        setattr(args, "xlat_from", xlat_from)
    else:
        # Ensure that there is an xlat field, even if it is only "None"
        log.debug("no language file specified")
        setattr(args, "xlat", None)
        setattr(args, "xlat_to", {})
        setattr(args, "xlat_from", {})

    return args


@entry_exit
def main(argv):
    """Mainline routine."""
    args = parse_args(argv)
    setup_logging(args)

    with sqlite3.connect(args.database) as con:
        try:
            cur = con.cursor()
            args.func(args, cur)

        except SystemExit as se:
            # This code ensures that we close down the DB on a sys.exit() call.
            return se.code

    # reach here and we succeeded.
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main(sys.argv))
